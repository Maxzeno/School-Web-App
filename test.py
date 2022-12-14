import pylightxl as pl

def write_excel(name: str, data: dict, sheet: str='sheet 1') -> bool:
	""" data = {'A1': {'v':10,'f':'','s':'', 'c': ''}, 'A2': {'v':20,'f':'','s':'', 'c': ''}} """
	def format_data(data):
		# formated_data1 = {}
		# for c, line in enumerate(data, 65):
		# 	for i, val in enumerate(line, 1):
		# 		formated_data1[f'{chr(c)}{i}'] = {'v': line[i-1],'f': '','s': '', 'c': ''}
		# # print(formated_data1)

		formated_data = {}
		A = 65
		for i in range(A, A+len(data[0])):
			for index, arr in enumerate(data, 1):
				formated_data[f'{chr(i)}{index}'] = {'v': arr[i-A],'f': '','s': '', 'c': ''}
		return formated_data

	# try:
	db = pl.Database()
	db.add_ws(sheet, format_data(data))
	res = pl.writexl(db, name)
	return True
	# except:
	# 	return False

def read_excel(path: str, sheet: str=None):
	""" Return 2D array -- inform of Generator"""
	book = pl.readxl(path)
	sheet = book.ws(sheet or book.ws_names[0])
	return sheet.rows

# from random import randint
# # data = {}
# # for c in range(65, 65+10):
# # 	for i in range(1, 11):
# # 		data[f'{chr(c)}{i}'] = {'v': ''.join([chr(randint(65, 65+26)) for i in range(randint(2, 10))]),'f': '','s': '', 'c': ''}

# ## data = {'A1': {'v':5,'f':'','s':'', 'c': ''}, 'A2': {'v':20,'f':'','s':'', 'c': ''}}
# data = [[2,3,4], [4,5,6], [5,6,7], [1,1,1]]

name = 'ssss (1).xlsx'
# # write_excel(name, data)

# print([i for i in read_excel(name)])







# import pyexcel as pe

# data = {'sheet 1': [[2,4,5], [4,5,6]]}


# data = pe.iget_records(file_name='grades.xlsx')
# # help(pe.get_records)
# print([i for i in data])

# # # book = pe.Book()
# # # print(dir(pe))









# import xls2xlsx as xl 

# xls = xl('student_mark_excel.xls')
# xlsx = xls.to_xlsx()
# print(xlsx)








# import excel21jsonapi

# excel21jsonapi.xlrd.open_workbook("grades (1).xlsx")



# from excel2json import convert_from_file
# import excel2json
# print(dir(excel2json))
# help(convert_from_file)
# data = convert_from_file('grades.xls')
# print(data)
# from excel import OpenExcel

# f = OpenExcel('excel.xlsx')
# print(dir(f))
# print(dir(f.data))
# quit()





# import xlrd
# book = xlrd.open_workbook("excel.xlsx")

# sh = book.sheet_by_index(0)

# # Iterate through rows, returning each as a list that you can index:
# for rownum in range(sh.nrows):
#     print(sh.row_values(rownum))

# print('\n emmm \n')
# # If you just want the first column:
# first_column = sh.row_values(0)
# print(first_column)


# print(dir(book))
# # print("The number of worksheets is {0}".format(book.nsheets))
# # print("Worksheet name(s): {0}".format(book.sheet_names()))
# sh = book.sheet_by_index(0)
# # print("{0} {1} {2}".format(sh.name, sh.nrows, sh.ncols))
# # print("Cell D30 is {0}".format(sh.cell_value(rowx=29, colx=3)))
# for rx in range(sh.nrows):
#     print(sh.row(rx))








from openpyxl import Workbook, load_workbook

book = load_workbook(name)
sheet = book.active

print(next(sheet.values))

for i in sheet.values:
	print(i)

# book = Workbook()
# sheet = book.active

# d = [list('abcd'), list('ghsd')]

# # for i, line in enumerate(d):
# # 	for j, val in enumerate(line):
# # 		print(val)
# # 		sheet.cell(row=i+1, column=j+1).value = val

# # print(dir(sheet))


# for i in d:
# 	sheet.append(i)

# for i in sheet.values:
# 	print(i)


# import xlrd

# print(dir(xlrd))

# help(xlrd.open_workbook)

# book = xlrd.open_workbook('aa.xls')
# print(dir(book))



# # import xlsxwriter module
# import xlsxwriter

# workbook = xlsxwriter.Workbook('Example3.xls')

# # By default worksheet names in the spreadsheet will be
# # Sheet1, Sheet2 etc., but we can also specify a name.
# worksheet = workbook.add_worksheet("My sheet")

# # Some data we want to write to the worksheet.
# scores = (
# 	['ankit', 1000],
# 	['rahul', 100],
# 	['priya', 300],
# 	['harshita', 50],
# )

# # Start from the first cell. Rows and
# # columns are zero indexed.
# row = 0
# col = 0

# # Iterate over the data and write it out row by row.
# for name, score in (scores):
# 	worksheet.write(row, col, name)
# 	worksheet.write(row, col + 1, score)
# 	row += 1

# workbook.close()


# from excel import OpenExcel
# f = OpenExcel('grades.xlsx')
# f = OpenExcel('grades.xlsx')
# f.read() # read all

# import pandas as pd
# data = pd.read_excel('grades.xlsx')
# print(data)





# # import xlsxwriter module
# import xlsxwriter

# # Workbook() takes one, non-optional, argument
# # which is the filename that we want to create.
# workbook = xlsxwriter.Workbook('hello.xlsx')

# # The workbook object is then used to add new
# # worksheet via the add_worksheet() method.
# worksheet = workbook.add_worksheet()

# # Use the worksheet object to write
# # data via the write() method.
# worksheet.write('A1', 'Hello..')
# worksheet.write('B1', 'Geeks')
# worksheet.write('C1', 'For')
# worksheet.write('D1', 'Geeks')

# # Finally, close the Excel file
# # via the close() method.
# workbook.close()



# import pandas as pd
# data = pd.read_excel('path/input.xlsx')
# print (data)




# from excel import OpenExcel
# f = OpenExcel('test.xls')
# f.read() # read all
# f.read('A') # read 'A' row
# f.read(1) # f.read('1'), read '1' column
# f.read('A5') # read 'A5' position
















# import requests
# from bs4 import BeautifulSoup as soup
# import os
# # import cssutils

# class PageScaper:
#     def __init__(self, url=None, login_url='https://cdssenugu.org/portal/login/validate_login', domain_url=None):
#         self.url = url
#         self.login_url = login_url
#         self.domain_url = domain_url or '/'.join(url.split('/')[:3])


#     def get_static_urls(self):
#         with requests.session() as request:
#             self.session_login(request)
#             resp = request.get(self.url).text
#             parse = soup(resp, 'lxml')

#             links = parse.find_all('link')
#             imgs = parse.find_all('img')
#             scripts = parse.find_all('script')
#             # as_ = parse.find_all('a')

#             urls = [self.base_urler(link.attrs.get('href')) for link in links if self.base_urler(link.attrs.get('href'))]
#             urls += [self.base_urler(script.attrs.get('src')) for script in scripts if self.base_urler(script.attrs.get('src'))]
#             urls += [self.base_urler(img.attrs.get('src')) for img in imgs if self.base_urler(img.attrs.get('src'))]
#             # urls += [self.base_urler(a_.attrs.get('href')) for a_ in as_ if self.base_urler(a_.attrs.get('href'))]

#         return set(urls)


#     def get_page_urls(self, url=None):
#         the_url = url if url else self.url
#         with requests.session() as request:
#             self.session_login(request)
#             resp = request.get(the_url).text
#             parse = soup(resp, 'lxml')

#             as_ = parse.find_all('a')
#             forms = parse.find_all('form')

#             urls = [self.base_urler(a_.attrs.get('href')) for a_ in as_ if self.base_urler(a_.attrs.get('href'))]
#             ## do not send any request to those wheather GET, POST etc it DELETES things urls = [self.base_urler(a_.attrs.get('onclick').split("('")[1].split("',")[0]) for a_ in as_ if self.base_urler(a_.attrs.get('onclick'))]
#             urls += [self.base_urler(form.attrs.get('action')) for form in forms if self.base_urler(form.attrs.get('action'))]
#             print(set(urls))
#         return set(urls)


#     def base_urler(self, url):
#         if url:
#             if url.startswith(self.domain_url):
#                 return url
#             elif url.startswith('http://') or url.startswith('https://'):
#                 return None
#             else:
#                 if url.startswith('/'):
#                     return self.domain_url + url
#                 else:
#                     return self.domain_url + "/" + url
#         return  None


#     def download_static(self, urls_list=None, folder=''):
#         folder += '/' if folder else ''
#         if urls_list:
#             self.session_login(request)
#             urls = urls_list  
#         else:
#             urls = self.get_static_urls()

#         with requests.session() as request:
#             for url in urls:
#                 path = folder + '/'.join(url.split('/')[3:-1])
#                 try:
#                     os.makedirs(path)
#                 except FileExistsError:
#                     pass
                    
#                 file_name = path+"/"+url.split('/')[-1]
#                 print(file_name)
#                 if not os.path.exists(file_name):
#                     print('downloading ... ', file_name)
#                     file = open(file_name, 'wb')
#                     got = request.get(url)
#                     file.write(got.content)
#                     file.close()

#         return 'Done'

#     def download_page(self, url=None, page_type='html', folder='', folder_in=''):
#         folder += '/' + folder_in if folder_in else folder_in
#         the_url = url if url else self.url
#         with requests.session() as request:
#             self.session_login(request)
#             path = folder
#             try:
#                 os.makedirs(path)
#             except FileExistsError:
#                 pass

#             d_appender = ''
#             for i in range(1, len(the_url.split('/'))+1):
#                 if not the_url.split('/')[i*-1].isdigit():
#                     name = the_url.split('/')[i*-1] + d_appender
#                     break
#                 else:
#                     d_appender = '_' + the_url.split('/')[i*-1] if not d_appender else '_' + the_url.split('/')[i*-1] + d_appender
#             else:
#                 name = d_appender[1:]

#             file_name = path+"/"+name+'.' + page_type
#             if not os.path.exists(file_name):
#                 file = open(file_name, 'wb')
#                 got = request.get(the_url)
#                 file.write(got.content)
#                 file.close()
#                 return 'Done'
#             return f'File with same name Exists in \'{file_name} \''


#     def session_login(self, session):
#         # got = session.get(url).content
#         # parse = soup(got, 'lxml')
#         # csrf_token = parse.find_all('input', attrs={'id':'csrf_token'})[0].attrs.get('value')

#         posted = session.post(self.login_url, 
#             data={'email':'Commandant@cdssenugu.org', 'password':'VPAdmin'})

#         html = posted.content
#         status = posted.status_code
#         return session, html, status

# urls = {'https://cdssenugu.org/portal/admin/management_team', 
# 'https://cdssenugu.org/portal/login/logout', 
# 'https://cdssenugu.org/portal/admin/noticeboard', 
# 'https://cdssenugu.org/portal/admin/accountant', 
# 'https://cdssenugu.org/portal/admin/librarian', 
# 'https://cdssenugu.org/portal/admin/dashboard', 
# 'https://cdssenugu.org/portal/admin/event_calendar', 
# 'https://cdssenugu.org/portal/', 
# 'https://cdssenugu.org/portal/admin/principal_comment_codes', 
# 'https://cdssenugu.org/portal/admin/manage_class', 
# 'https://cdssenugu.org/portal/admin/pass_mark', 
# 'https://cdssenugu.org/portal/admin/profile', 
# 'https://cdssenugu.org/portal/admin/tabulation_sheet', 
# 'https://cdssenugu.org/portal/admin/invoice', 
# 'https://cdssenugu.org/portal/admin/cognitive_key_domain_scores', 
# 'https://cdssenugu.org/portal/admin/permission', 
# 'https://cdssenugu.org/portal/admin/promotion', 
# 'https://cdssenugu.org/portal/admin/expense_category', 
# 'https://cdssenugu.org/portal/admin/book', 
# 'https://cdssenugu.org/portal/admin/sb_description_score_manage', 
# 'https://cdssenugu.org/portal/admin/housemaster_comment_codes', 
# 'https://cdssenugu.org/portal/admin/student', 
# 'https://cdssenugu.org/portal/admin/class_room', 
# 'https://cdssenugu.org/portal/admin/teacher', 
# 'https://cdssenugu.org/portal/admin/expense', 
# 'https://cdssenugu.org/portal/admin/attendance', 
# 'https://cdssenugu.org/portal/admin/teacher_comment_codes', 
# 'https://cdssenugu.org/portal/admin/principal_subject_codes', 
# 'https://cdssenugu.org/portal/admin/book_issue', 
# 'https://cdssenugu.org/portal/admin/student/create', 
# 'https://cdssenugu.org/portal/admin/routine', 
# 'https://cdssenugu.org/portal/admin/mark', 
# 'https://cdssenugu.org/portal/admin/grade', 
# 'https://cdssenugu.org/portal/admin/subject', 
# 'https://cdssenugu.org/portal/admin/subject_permission', 
# 'https://cdssenugu.org/portal/admin/exam', 
# 'https://cdssenugu.org/portal/admin/syllabus', 
# 'https://cdssenugu.org/portal/admin/school_settings', 
# 'https://cdssenugu.org/portal/admin/parent', 
# 'https://cdssenugu.org/portal/admin/sb_description', 
# 'https://cdssenugu.org/portal/admin/department'}

# page_scraper = PageScaper('https://cdssenugu.org/portal/admin/department')

# # print(page_scraper.download_static(folder='static_store/dashboard'))
# # print(page_scraper.download_page(folder='static_store', folder_in='new'))
# # print(page_scraper.get_static_urls())



# lst_of_set = [page_scraper.get_page_urls(url) for url in  urls]
# print(lst_of_set)

# def unique_except_set(lst_of_set, except_url_list):
#     new_set = set()
#     def unique_x_url(set_iter, except_url_list):
#         return set_iter.difference(except_url_list)
#     for set_ in  lst_of_set:
#         new_set.update(set_)

#     return unique_x_url(new_set, except_url_list)

# for i in unique_except_set(lst_of_set, urls):
#     page_scraper.download_page(i, folder='static_store', folder_in='new')

# # resp = requests.get('http://127.0.0.1:8000/portal/admin/teacher').text
# # parse = soup(resp, 'lxml')

# # ds = parse.find_all('a')
# # urls = [d.attrs.get('onclick') for d in ds if d.attrs.get('onclick')]
# # for url in urls[:2]:
# #     print(url.split("('")[1].split("',")[0])
# #     page_scraper.download_page(url.split("('")[1].split("',")[0], folder='static_store', folder_in='new1')
# #     